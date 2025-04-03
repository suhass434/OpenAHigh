import fitz  # This is PyMuPDF
import re
import os
import logging
import time
import json
import argparse
from pathlib import Path
from dotenv import load_dotenv
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Initialize logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configure Gemini API
load_dotenv()
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
if not GEMINI_API_KEY:
    logger.warning("GEMINI_API_KEY not found in environment variables")
else:
    genai.configure(api_key=GEMINI_API_KEY)

# Style configurations
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

DOCUMENT_MAP = {}
# DOCUMENT_MAP = {
#     "BW2024-08": {
#         "document_id": "BW2024-08",
#         "affected_product": "Experion PKS",
#         "affected_assets": "UOC\nEIM\nELCN Bridge\nELCN Node",
#         "affected_release": "R511.5 up to R511.5 TCU5 HF2\nR520.2 up to R520.2 TCU6 HF3\nR530",
#         "fixed_release": "R511.5 TCU6 and later\nR520.2 TCU7 and later\nR530 TCU1 and later",
#         "par_number": "1-G9ENCXT",
#         "configuration": "All Configurations"
#     },
#     "PN2024-07A": {
#         "document_id": "PN2024-07A",
#         "affected_product": "Experion PKS",
#         "affected_assets": "C300 CC-PCNT05\nCN100\nUOC\nEIM\nC300PM\nEHPMX\nUEA\nUVA",
#         "affected_release": "All Releases",
#         "fixed_release": "R511.5 TCU6 and later\nR520.2 TCU6 HF2 and later\nR520.2 TCU7 and later",
#         "par_number": "1-FM396Q5\nREUCN-7883",
#         "configuration": "Redundant assets"
#     },
# }

def extract_document_id(pdf_path):
    """Extract document ID from filename with validation"""
    try:
        match = re.search(r'\b(BW|PN)\d{4}-\d+[A-Z]?\b', Path(pdf_path).stem.upper())
        print(match)
        return match.group(0) if match else "UNKNOWN"
    except Exception as e:
        logger.error(f"Error extracting document ID: {str(e)}")
        return "UNKNOWN"

def extract_with_llm(text, doc_id):
    """Enhanced LLM extraction with sliding window"""
    chunk_size = 12000  # Keep original truncation size
    overlap = 3000  # Preserve context between chunks
    max_chunks = 5  # Prevent infinite processing
    base_delay = 15
    max_retries = 3
    required_keys = ["affected_assets", "affected_release", 
                    "fixed_release", "par_number", "configuration"]
    
    combined_result = {}
    
    for chunk_idx in range(max_chunks):
        start = chunk_idx * (chunk_size - overlap)
        end = start + chunk_size
        chunk_text = text[start:end]
        
        if not chunk_text:
            break  # End of document

        prompt = f"""Analyze this technical document and extract the following information:
        - Affected assets (list of hardware/software components)
        - Affected software releases/versions
        - Fixed releases/versions with updates
        - PAR numbers (problem report IDs)
        - Affected configurations

        Format response as JSON with these keys:
        "affected_assets", "affected_release", "fixed_release", "par_number", "configuration"

        Example response for PN2024-08:
        {{
            "affected_assets": "C300 CC-PCNT05\\nC300 CC-PCNT02\\nC300 CC-PCNT01\\nUOC\\nC200E\\nC300PM",
            "affected_release": "R501.x\\nR510.x\\nR511.x\\nR520.x\\nR530",
            "fixed_release": "R510.2 Hotfix 15 and later\\nR511.5 TCU6 and later\\nR520.2 TCU7 and later\\nR530 TCU1 and later",
            "par_number": "1-FWPPZNL",
            "configuration": "Batch Configuration with Recipe Control Module (RCM)"
        }}

        Document text:
        {chunk_text}
        """

        for attempt in range(max_retries):
            try:
                model = genai.GenerativeModel('gemini-1.5-flash')
                response = model.generate_content(prompt)
                json_str = response.text.replace('```json', '').replace('```', '').strip()
                chunk_data = json.loads(json_str)

                # Merge new findings without overwriting existing values
                for key in required_keys:
                    if key not in combined_result or not combined_result[key]:
                        combined_result[key] = chunk_data.get(key, "")

                # Early exit if we've found all fields
                if all(combined_result.get(k) for k in required_keys):
                    return combined_result

                break  # Successfully processed chunk

            except json.JSONDecodeError:
                logger.warning(f"Invalid JSON in chunk {chunk_idx+1}, attempt {attempt+1}")
                if attempt == max_retries-1:
                    continue  # Move to next chunk after final retry
            except Exception as e:
                logger.error(f"Chunk {chunk_idx+1} failed: {str(e)}")
                if '429' in str(e):
                    time.sleep(base_delay * (2 ** attempt))
                continue

        # Early exit if all fields found
        if all(combined_result.get(k) for k in required_keys):
            break

    return combined_result if any(combined_result.values()) else None

def extract_sections(text, doc_id):
    """Enhanced hybrid extraction with improved patterns"""
    sections = {
        "document_id": doc_id,
        "affected_product": "Experion PKS",
        "affected_assets": "",
        "affected_release": "",
        "fixed_release": "",
        "par_number": "",
        "configuration": ""
    }

    # Enhanced regex patterns
    try:
        # Assets with multi-line support
        assets_match = re.search(
            r'(?:AFFECTED COMPONENTS?|PRODUCT AFFECTED):\s*((?:(?:C300|UOC|EIM|ELCN|C200E?|C300PM|EHPMX|UEA|UVA|CN100|Experion PKS Servers)[\s\S]*?))(?:\.|$|\n\S+)',
            text, re.IGNORECASE)
        if assets_match:
            sections["affected_assets"] = "\n".join(
                [a.strip() for a in re.split(r', |\n|;', assets_match.group(1)) if a.strip()]
            )

        # Releases with version range support
        release_match = re.search(
            r'(?:AFFECTED RELEASE|VERSION AFFECTED):\s*((?:(?:R\d+\.\d+[^\n]*)(?:\n|\\n|, |; | or ))+)',
            text, re.IGNORECASE)
        if release_match:
            sections["affected_release"] = "\n".join(
                [r.strip() for r in re.split(r', |\n|;| or ', release_match.group(1)) if r.strip()]
            )

        # Fixed releases with update details
        fixed_match = re.search(
            r'(?:FIXED IN|RESOLVED IN):\s*((?:(?:R\d+\.\d+[^\n]*)(?:\n|\\n|, |; | and ))+)',
            text, re.IGNORECASE)
        if fixed_match:
            sections["fixed_release"] = "\n".join(
                [f.strip() for f in re.split(r', |\n|;| and ', fixed_match.group(1)) if f.strip()]
            )

        # PAR numbers with multi-value support
        par_match = re.finditer(
            r'(?:PAR|PROBLEM REPORT) (?:NUMBERS?|ID):?\s*([A-Z0-9-]+(?:\s*,\s*[A-Z0-9-]+)*)',
            text, re.IGNORECASE)
        if par_match:
            sections["par_number"] = "\n".join(
                [m.group(1).replace(',', '\n') for m in par_match]
            )

        # Configuration context
        config_match = re.search(
            r'(?:CONFIGURATION AFFECTED|AFFECTED SETUP):\s*(.+?)(?=\n\S+:|$)', 
            text, re.IGNORECASE | re.DOTALL)
        if config_match:
            sections["configuration"] = config_match.group(1).strip()

    except Exception as e:
        logger.error(f"Regex extraction error: {str(e)}")

    # LLM augmentation for all fields
    llm_data = extract_with_llm(text, doc_id)
    if llm_data:
        for key in sections:
            # Only overwrite if LLM provides better data
            current_value = sections[key].strip()
            llm_value = llm_data.get(key, "").strip()
            
            if key == "affected_assets":
                if len(llm_value.split('\n')) > len(current_value.split('\n')):
                    sections[key] = llm_value
            elif key in ["affected_release", "fixed_release"]:
                if any(x in llm_value.lower() for x in ["hotfix", "tcu", "hf"]) and not any(x in current_value.lower() for x in ["hotfix", "tcu", "hf"]):
                    sections[key] = llm_value
            elif not current_value and llm_value:
                sections[key] = llm_value

    # Post-processing cleanup
    sections["par_number"] = sections["par_number"].rstrip(',')
    sections["configuration"] = sections["configuration"].replace('Confoguration', 'Configuration')
    
    return sections

def process_pdf(pdf_path):
    """Processing with enhanced validation"""
    try:
        # Convert to Path object if string
        pdf_path = Path(pdf_path) if isinstance(pdf_path, str) else pdf_path
        
        if not pdf_path.exists():
            logger.error(f"PDF file not found: {pdf_path}")
            return None
            
        doc = fitz.open(str(pdf_path))
        text = " ".join([page.get_text("text") for page in doc])
        doc_id = extract_document_id(str(pdf_path))
        
        if doc_id in DOCUMENT_MAP:
            logger.info(f"Using predefined mapping for {doc_id}")
            return DOCUMENT_MAP[doc_id]
            
        result = extract_sections(text, doc_id)
        
        # Final validation
        required_fields = ["affected_assets", "affected_release", "fixed_release"]
        if not all(result.get(field) for field in required_fields):
            logger.warning(f"Incomplete data for {doc_id}, retrying with full LLM")
            llm_data = extract_with_llm(text, doc_id)
            if llm_data:
                result.update(llm_data)
        
        return result
        
    except Exception as e:
        logger.error(f"Error processing {pdf_path}: {str(e)}")
        return None

def create_excel_output(metadata_list, output_path):
    """Create Excel output with structured formatting and error handling"""
    try:
        # Convert to Path object if string
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata"

        # ===== Style Definitions =====
        GREEN_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FF000000")
        TITLE_FONT = Font(bold=True, size=14)
        WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top", horizontal="left")
        CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
        THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # ===== Main Header =====
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(metadata_list)+2)
        main_header = ws.cell(row=1, column=1, value="Extracted Metadata")
        main_header.font = TITLE_FONT
        main_header.alignment = CENTER_ALIGNMENT
        main_header.border = THIN_BORDER

        # ===== Header Row =====
        ws.cell(row=2, column=1, value="")  # Empty first column
        field_header = ws.cell(row=2, column=2, value="Metadata Fields")
        field_header.font = HEADER_FONT
        field_header.alignment = CENTER_ALIGNMENT
        field_header.border = THIN_BORDER

        # Document ID headers
        for col_idx, metadata in enumerate(metadata_list, start=3):
            cell = ws.cell(row=2, column=col_idx, value=metadata.get("document_id", "UNKNOWN"))
            cell.fill = GREEN_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER

        # ===== Data Rows =====
        fields = [
            ("affected_product", "Affected Product"),
            ("affected_assets", "Affected Assets"),
            ("affected_release", "Affected Release"),
            ("fixed_release", "Fixed Release"),
            ("par_number", "PAR Number"),
            ("configuration", "Affected Configuration")
        ]

        for row_idx, (field_key, field_label) in enumerate(fields, start=3):
            # Field label
            label_cell = ws.cell(row=row_idx, column=2, value=field_label)
            label_cell.font = HEADER_FONT
            label_cell.alignment = WRAP_ALIGNMENT
            label_cell.border = THIN_BORDER

            # Field values
            for col_idx, metadata in enumerate(metadata_list, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=metadata.get(field_key, ""))
                cell.fill = GREEN_FILL
                cell.alignment = WRAP_ALIGNMENT
                cell.border = THIN_BORDER

        # ===== Column Sizing =====
        ws.column_dimensions['A'].width = 5  # Empty first column
        ws.column_dimensions['B'].width = 25  # Field labels
        for col in range(3, len(metadata_list) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 30

        # Save the workbook
        wb.save(str(output_path))
        logger.info(f"Successfully created Excel file: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Excel creation error: {str(e)}")
        raise  # Re-raise the exception to be handled by the caller

def main():
    parser = argparse.ArgumentParser(description='Process SCN documents.')
    parser.add_argument('--input', required=True, help='Input directory containing PDFs')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    args = parser.parse_args()

    pdf_files = [f for f in Path(args.input).glob("*.pdf") if f.is_file()]
    metadata = []
    
    for pdf in pdf_files:
        logger.info(f"Processing: {pdf.name}")
        result = process_pdf(pdf)
        if result:
            metadata.append(result)
            logger.info(f"Successfully extracted metadata from: {pdf.name}")
    
    if metadata:
        if create_excel_output(metadata, args.output):
            logger.info(f"Processing completed. Output saved to: {args.output}")
        else:
            logger.error("Failed to create Excel file")

if __name__ == "__main__":
    main()