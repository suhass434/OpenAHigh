import pandas as pd
import os
import json
from dotenv import load_dotenv
import google.generativeai as genai
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Load environment variables
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")

# Configure Gemini API
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

# Load compatibility matrix
COMPATIBILITY_MATRIX = pd.read_excel(
    os.path.join(os.path.dirname(__file__), "../", "data", "Compatible Versions.xlsx"),
    sheet_name="Sheet1"
)

def get_latest_version(installed_versions_check, compatibility_matrix):
    """Get latest compatible versions using Gemini AI."""
    prompt = f"""
    COMPARE THE MIGRATION VERSIONS ACCORDING TO THE BELOW RULES:
    - OLD INSTALLED VERSION:
    {json.dumps(installed_versions_check, indent=2)}

    - COMPATIBILITY MATRIX WITH VERSIONS FOR THE NEW VERSION:
    {json.dumps(compatibility_matrix, indent=2)}
    
    - Always choose the latest for Domain Controller: 
    If INSTALLED version = Windows Server 2019, TO RECOMMENDED VERSION = Windows Server 2022 
    - Always choose the latest supported for Server Hardware: 
    INSTALLED version = Dell PE R450 / XR11 Server, TO RECOMMENDED VERSION = Dell R740XL Server
    - Always choose the latest compatible for Workstation Hardware: 
    INSTALLED version = Dell R7910XL Workstation, TO RECOMMENDED VERSION = Dell R7910XL Workstation 

    OUTPUT FORMAT:
    {{
    "Domain Controller": RECOMMENDED VERSION FOR Domain Controller,
    "Server Hardware": RECOMMENDED VERSION FOR Server Hardware,
    "Workstation Hardware": RECOMMENDED VERSION FOR Workstation Hardware
    }}
    """

    try:
        response = model.generate_content(prompt)
        cleaned_response = response.text.strip().strip("```json").strip("```")
        return json.loads(cleaned_response)
    except Exception as e:
        print(f"Error in Gemini API call: {str(e)}")
        # Fallback default values
        return {
            "Domain Controller": "Windows Server 2022",
            "Server Hardware": "Dell R740XL Server",
            "Workstation Hardware": "Dell R7910XL Workstation"
        }

def process_migration(installed_source: str, installed_versions: dict, target_version: str) -> dict:
    """Process migration and return recommended versions."""
    recommended_versions = {}
    
    try:
        supported_versions = COMPATIBILITY_MATRIX[COMPATIBILITY_MATRIX["Experion PKS"] == target_version]
        
        installed_versions_check = {
            "Domain Controller": installed_versions.get("Domain Controller", ""),
            "Server Hardware": installed_versions.get("Server Hardware", ""),
            "Workstation Hardware": installed_versions.get("Workstation Hardware", "")
        }
        
        compatibility_matrix_for_llm = {}
        for software in ["Domain Controller", "Server Hardware", "Workstation Hardware"]:
            available_versions = supported_versions[software].dropna().tolist()
            compatibility_matrix_for_llm[software] = available_versions
        
        # Get LLM recommendations
        llm_recommendations = get_latest_version(installed_versions_check, compatibility_matrix_for_llm)
        
        # Build final recommendations
        for software, version in installed_versions.items():
            if software in ["Domain Controller", "Server Hardware", "Workstation Hardware"]:
                recommended_versions[software] = llm_recommendations[software]
            else:
                recommended_versions[software] = version
        
        return recommended_versions
        
    except Exception as e:
        raise Exception(f"Error processing migration: {str(e)}")

def generate_migration_excel(installed_source: str, installed_versions: dict, target_version: str) -> BytesIO:
    """Generate Excel file with migration results."""
    try:
        # Get recommended versions
        recommended_versions = process_migration(installed_source, installed_versions, target_version)
        
        # Create lists for DataFrame
        software_list = list(installed_versions.keys())
        installed_list = [installed_versions[software] for software in software_list]
        recommended_list = [recommended_versions[software] for software in software_list]
        
        # Create DataFrame
        migration_df = pd.DataFrame({
            "Installed Software": software_list,
            "Installed Versions": installed_list,
            "Recommended Versions": recommended_list
        })
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        
        # Add header
        ws.merge_cells('A1:C1')
        header_cell = ws['A1']
        header_cell.value = f"Experion Migration Target: {target_version}"
        header_cell.font = Font(bold=True, size=14)
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(migration_df, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply styling
        sky_blue = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
        light_green = PatternFill(start_color='85E085', end_color='85E085', fill_type='solid')
        header_yellow = PatternFill(start_color='FFFF80', end_color='FFFF80', fill_type='solid')
        
        # Style header row
        for cell in ws[2]:
            cell.fill = header_yellow
        
        # Style data columns
        for row in ws.iter_rows(min_row=3, max_col=2):
            for cell in row:
                cell.fill = sky_blue
                
        for row in ws.iter_rows(min_row=3, min_col=3, max_col=3):
            for cell in row:
                cell.fill = light_green
        
        # Save to BytesIO
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return excel_bytes
        
    except Exception as e:
        raise Exception(f"Error generating Excel file: {str(e)}") 