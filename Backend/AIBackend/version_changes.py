# IMPORTS
import pandas as pd
import os
import json
from dotenv import load_dotenv
import google.generativeai as genai


# Load API Key from .env file
load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY")


# Configure Gemini API
genai.configure(api_key=API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')


# I/O paths for the CSV files
input_csv = r"C:\Users\shama\Desktop\HoneyWell Hackathon\Task 2\Task2 Dataset\Input\Compatibile Versions.xlsx"
output_csv = r"C:\Users\shama\Desktop\HoneyWell Hackathon\Task 2\Test Outputs\migration_output2.xlsx"

# LLM Function to get latest versions (can be later changed with better quality prompts after breifings)
def get_latest_version(installed_versions_check, compatibility_matrix):
    # Format Prompt
    prompt = f"""
    COMPARE THE MIGRATION VERSIONS ACCORDING TO THE BELOW RULES:
    - OLD INSTALLED VERSION:
    {json.dumps(installed_versions_check, indent=2)}

    - COMPATIBILITY MATRIX WITH VERSIONS FOR THE NEW VERSION:
    {json.dumps(compatibility_matrix, indent=2)}
    
    - Always choose the latest for Domain Controller	: 
    If INSTALLED version = Windows Server 2019, , TO RECOMMENDED VERSION = 
    Windows Server 2022 
    - Always choose the latest supported for Server Hardware: 
    INSTALLED version = Dell PE R450 / XR11 Server, 
    TO RECOMMENDED VERSION = Dell R740XL Server
    - Always choose the latest compatible for Workstation Hardware: SRC: 
    INSTALLED version = Dell R7910XL Workstation,
    TO RECOMMENDED VERSION = Dell R7910XL Workstation 

    OUTPUT FORMAT:
    {{
    "Domain Controller": RECOMMENDED VERSION FOR Domain Controller,
    "Server Hardware": RECOMMENDED VERSION FOR Server Hardware,
    "Workstation Hardware": RECOMMENDED VERSION FOR Workstation Hardware
    }}

    Example:
    "Domain Controller": "Windows Server 2022",
    "Server Hardware": "Dell R740XL Server",
    "Workstation Hardware": "Dell R7910XL Workstation"
    """

    # Generate response using Gemini API
    response = model.generate_content("Hello, can you respond?")
    print(response.text)


    response = model.generate_content(prompt)
    print("Raw Gemini Response:", response.text)  # Debugging step

    try:
        # Parse JSON safely
        cleaned_response = response.text.strip().strip("```json").strip("```")
        print("Cleaned Gemini Response:", cleaned_response)  # Debugging step
        recommended_versions = json.loads(cleaned_response)
    
        return dict(recommended_versions)



    except json.JSONDecodeError:
        print("Error: Failed to parse Gemini response as JSON.")
        return []
            
    

# Processing/Analysing Functions
def migrate_versions(installed_versions, compatibility_csv, versions_target):
    """
    Function to apply migration rules and determine recommended versions.
    """
    recommended_versions = {}
    
    try: 
        supported_versions = compatibility_csv[compatibility_csv["Experion PKS"] == versions_target]
    except KeyError:
        print(f"Error: The specified version {versions_target} is not found in the compatibility matrix. Please enter the correct version of migration target.")
        exit()
    
    installed_versions_check = {"Domain Controller": installed_versions["Domain Controller"], "Server Hardware": installed_versions["Server Hardware"], "Workstation Hardware": installed_versions["Workstation Hardware"]}
    
    compatibility_matrix_for_llm = {}
    
    for software, version in installed_versions.items():
        # keep the same version if supported
        if software in [
            "Blending and Movement", "Control Performane Monitor", "Dynamo", 
            "Dynamo Operations Suite (DOS)", "Field Device Manager (FDM)", "Operating System", "Profit Suite"
        ]:
            recommended_versions[software] = version  
        elif software == "Domain Controller" or software == "Server Hardware" or software == "Workstation Hardware":
            # making a list of versions from the compatibility matrix
            
            available_versions = supported_versions[software].dropna().tolist()
            compatibility_matrix_for_llm[software] = available_versions
        
        else:
            # default case (keep same version)
            recommended_versions[software] = version  
    
    llm_recommendations = get_latest_version(installed_versions_check, compatibility_matrix_for_llm)
    recommended_versions["Domain Controller"] = llm_recommendations["Domain Controller"]
    recommended_versions["Server Hardware"] = llm_recommendations["Server Hardware"]
    recommended_versions["Workstation Hardware"] = llm_recommendations["Workstation Hardware"]

    recommended_versions["Experion PKS"] = versions_target  
    return recommended_versions



# USER INPUTS

# 0. INSTALLED SOFTWARE

# Define installed software and versions

installed_software = [
    "Experion PKS", "Blending and Movement", "Control Performane Monitor", "Domain Controller",
    "Dynamo", "Dynamo Operations Suite (DOS)", "Field Device Manager (FDM)",
    "Operating System", "Profit Suite", "Server Hardware", "Workstation Hardware"
]

installed_versions = {
    "Experion PKS": "R520.1",
    "Blending and Movement": "FBM R520.1",
    "Control Performane Monitor": "Taiji PID R320.1",
    "Domain Controller": "Windows Server 2019",
    "Dynamo": "M&R-R210.1",
    "Dynamo Operations Suite (DOS)": "Dyn Ops Monitoring-R220.1",
    "Field Device Manager (FDM)": "FDM R511.5",
    "Operating System": "Microsoft Windows Server 2019 Standard",
    "Profit Suite": "Profit Suite R512.1",
    "Server Hardware": "Dell PE R450 / XR11 Server",
    "Workstation Hardware": "Dell R7910XL Workstation"
}

# 2. INSTALLED EXPERION SOURCE
version_src = input("Enter the INSTALLED EXPERION SOURCE: (EX: one of the following, R520.2, R520.1, R530.1): ")

installed_versions["Experion PKS"] = version_src

# 3. INSTALLED Experion VERSION DETAILS
print(f"Enter DETAILS for the INSTALLED EXPERION SOURCE: {installed_versions['Experion PKS']}")
for key, value in installed_versions.items():
    if(key == "Experion PKS"):
        continue
    installed_versions[key] = input(f" VERSION/TYPE DETAILS FOR: {key}: ") or installed_versions[key]

    
# OR COLLECT DETAILS FROM A CSV FILE or default hardcoded values
# installed_versions = pd.read_csv(input_csv, sheet_name="Installed Versions")


# 3. RECOMMENDED Experion Migration Target
versions_target = input("Enter the RECOMMENDED Experion Migration Target in CAPITALS: (EX: one of the following, R520.2, R520.1, R530.1): ")

# 4. EXTRACTING RECOMMENDED VERSION DETAILS
print(f"DETAILS for the RECOMMENDED Experion Migration Target: {versions_target}")
compatibility_csv = pd.read_excel(input_csv, sheet_name="Sheet1")  


# 5. COMPARE THE INSTALLED VERSION WITH THE RECOMMENDED VERSION according to the compatibility matrix

# Get recommended versions
recommended_versions = migrate_versions(installed_versions, compatibility_csv, versions_target)
recommended_versions_list = [recommended_versions.get(software, "N/A") for software in installed_software]

# 6. FINAL WRAPS

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


# Create a DataFrame
migration_df = pd.DataFrame({
    "Installed Software": installed_software,
    "Installed Versions": list(installed_versions.values()),
    "Recommended Versions": recommended_versions_list
})


# Create a workbook and worksheet
wb = Workbook()
ws = wb.active

# Add the custom header
ws.merge_cells('A1:C1')
header_cell = ws['A1']
header_cell.value = f"Experion Migration Target: {versions_target}"

header_cell.font = Font(bold=True, size=14)

# Write the DataFrame starting from row 2
for r_idx, row in enumerate(dataframe_to_rows(migration_df, index=False, header=True), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Define colors
sky_blue = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
light_green = PatternFill(start_color='85E085', end_color='85E085', fill_type='solid')

# Apply colors to columns
for cell in ws[2]:  # Header row
    cell.fill = PatternFill(start_color='FFFF80', end_color='FFFF80', fill_type='solid')  # Light gray for header

for row in ws.iter_rows(min_row=3, max_col=2):  # First two columns (A and B)
    for cell in row:
        cell.fill = sky_blue

for row in ws.iter_rows(min_row=3, min_col=3, max_col=3):  # Third column (C)
    for cell in row:
        cell.fill = light_green

# Save the workbook
wb.save(output_csv)
# Save to CSV

# migration_df.to_csv(output_csv, index=False)

print(f"Migration rules applied. Output saved as {output_csv}.")
